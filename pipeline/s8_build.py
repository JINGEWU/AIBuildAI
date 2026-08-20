# -*- coding: utf-8 -*-
"""S8 交付构建（无批次）：把当前全部合格论文出成一套 PDF 目录 + 一张汇总表。

为什么不再分批：逼出分批的从来不是筛选逻辑，而是文件命名。
`A001_标题.pdf` 这种按复刻分排序的序号，一加新论文排名就变、所有文件都要重命名，
所以只能把每次的结果冻成一批。这里改用**与排序无关的稳定命名**——直接沿用
fulltext/ 的 `期刊_年_标题_DOI尾号.pdf`：加论文只新增文件，永不改名，
一个平铺目录就够，批次概念自然消失。排序交给表里的「复刻分」列。

幂等：只补缺失的文件、只删已不合格的，反复跑结果一致。

早期那 54 篇不再是特例：它们的抽取结果已由 `run.py info import` 导进 extracted.json，
和其他论文走同一条路，所以这里没有 --include-baseline 之类的开关。

用法: python3 run.py build [--out delivery] [--tiers SAB] [--xlsx-only] [--prune] [--dry-run]
"""
import os, sys, re, json, shutil, argparse, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config as C, common as m
from s9_enrich import domain, art_type, sota_status, REVIEW

DATA_RE = re.compile(r"(huggingface\.co/datasets|zenodo\.org|figshare\.com|kaggle\.com|"
                     r"physionet\.org|osf\.io|datadryad\.org|paperswithcode\.com)", re.I)
CODE_RE = re.compile(r"(github\.com|gitlab\.com|codeocean\.com)", re.I)
COLS = C.EXTRACT_COLUMNS + ["JCR IF", "期刊领域", "任务领域", "文章类型", "SOTA可对比性",
                            "复刻优先级", "复刻分", "基准公开", "来源", "PDF 文件名"]


def select(tiers):
    """当前全部合格论文。判据只有一套，与什么时候下载的无关。"""
    V = json.load(open(f"{C.DATA}/verdicts2.json", encoding="utf-8"))
    E = json.load(open(f"{C.DATA}/extracted.json", encoding="utf-8"))
    D = {m.norm_doi(r["doi"]): r for r in m.read_jsonl(f"{C.DATA}/fulltext_digest.jsonl", True)}
    st = m.load_shards("fetch_state")

    def xml_only(d):
        s = st.get(d) or {}
        return s.get("src") == "epmc-xml" or (s.get("file") or "").lower().endswith(".xml")

    rows, drop = [], collections.Counter()
    for d, v in V.items():
        if v.get("v") != "入选":
            drop["未入选"] += 1; continue
        if d not in E:
            drop["未抽取"] += 1; continue
        g = D.get(d, {})
        sc = m.repro_score(v, g, g.get("journal"))
        t = m.repro_tier(sc)
        if t not in tiers:
            drop[f"档位 {t}"] += 1; continue
        if d in REVIEW:
            drop["综述/复现报告"] += 1; continue
        if xml_only(d):
            drop["仅 XML 全文"] += 1; continue
        rows.append({"doi": d, "v": v, "g": g, "e": E[d], "score": sc, "tier": t})
    rows.sort(key=lambda x: -x["score"])
    return rows, drop


def main(argv):
    ap = argparse.ArgumentParser(prog="build")
    ap.add_argument("--out", default="delivery", help="交付目录名")
    ap.add_argument("--tiers", default="SAB", help="纳入的复刻档位")
    ap.add_argument("--xlsx-only", action="store_true", help="只重出表格，不碰 PDF")
    ap.add_argument("--prune", action="store_true", help="删掉目录里已不合格的 PDF（默认只报告）")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args(argv)

    OUT = a.out if os.path.isabs(a.out) else f"{C.ROOT}/{a.out}"
    PDFDIR = f"{OUT}/pdfs"
    rows, drop = select(tuple(a.tiers))
    print(f"合格 {len(rows)} 篇（档位 {a.tiers}）；排除: {dict(drop)}")
    if not rows:
        m.die("没有合格论文，检查 verdicts2.json / extracted.json")

    # 稳定命名：按 期刊_年_标题_DOI尾号 规范化，与排序无关。
    # 全文库的文件名跨时期不一致（早期用 IEEETrans_PatternAnaly、年份缺失记作 NA），
    # 交付时统一重算一次；DOI 尾号保证唯一，元数据齐了以后这个名字就不再变。
    files = m.file_index(("fulltext", "fulltext_xml", "papers", "arxiv_隔离"), with_origin=True)
    st = m.load_shards("fetch_state")
    missing = []
    for r in rows:
        s = st.get(r["doi"]) or {}
        hit = files.get(s.get("file") or "")
        if not hit:
            tail = r["doi"].split("/")[-1]
            hit = next((v for k, v in files.items() if len(tail) > 6 and tail[:20] in k), None)
        if not hit:
            missing.append(r); r["pdf"] = ""; r["origin"] = ""; continue
        r["src_path"], r["origin"] = hit
        g = r["g"]
        rec = {"doi": r["doi"], "journal": g.get("journal") or s.get("journal"),
               "year": g.get("year") or s.get("year"), "title": g.get("title") or s.get("title")}
        ext = os.path.splitext(hit[0])[1].lower()
        r["pdf"] = (m.fname(rec) + ext) if rec["title"] else os.path.basename(hit[0])
    keep = {r["pdf"] for r in rows if r["pdf"]}
    if len(keep) != len([r for r in rows if r["pdf"]]):
        dup = [k for k, n in collections.Counter(r["pdf"] for r in rows if r["pdf"]).items() if n > 1]
        m.die(f"命名冲突 {len(dup)} 组，交付会互相覆盖: {dup[:5]}")

    have = set(os.listdir(PDFDIR)) if os.path.isdir(PDFDIR) else set()
    have = {f for f in have if not f.startswith(".")}
    add, stale = keep - have, have - keep
    print(f"  PDF: 目录现有 {len(have)}，应有 {len(keep)}；待新增 {len(add)}，已不合格 {len(stale)}")
    if missing:
        print(f"  ⚠️ 缺全文 {len(missing)} 篇（未纳入 PDF，仍列在表里）:")
        for r in missing[:8]:
            print(f"     [{r['tier']}] {r['doi']}  {(r['g'].get('title') or '')[:52]}")
    if stale:
        for f in sorted(stale)[:8]:
            print(f"     - 多余: {f[:74]}")
        if not a.prune:
            print("     （加 --prune 才会删除）")

    if not (a.xlsx_only or a.dry_run):
        os.makedirs(PDFDIR, exist_ok=True)
        for r in rows:
            if r["pdf"] in add:
                shutil.copyfile(r["src_path"], f"{PDFDIR}/{r['pdf']}")
        if a.prune:
            for f in stale:
                os.remove(f"{PDFDIR}/{f}")
        n = len([f for f in os.listdir(PDFDIR) if not f.startswith(".")])
        print(f"  → {PDFDIR}  {n} 个文件")
    print("  档位:", dict(collections.Counter(r["tier"] for r in rows)))
    print("  来源:", dict(collections.Counter(r.get("origin") or "缺文件" for r in rows)))
    if a.dry_run:
        m.log("dry-run，未落盘"); return

    # ── 汇总表 ──
    _J = {j["name"]: j for j in C.journals()}
    from openpyxl.styles import Font, PatternFill
    wb = m.new_book(); ws = wb.create_sheet("汇总")
    ws.append(COLS)
    for i, r in enumerate(rows, 1):
        g, v, e = r["g"], r["v"], r["e"]
        lk = [x for x in (g.get("links") or []) if not re.match(r"openreview\.net", x, re.I)]
        dl = [x for x in lk if DATA_RE.search(x)]
        cl = [x for x in lk if CODE_RE.search(x)] + \
             [x for x in lk if re.match(r"huggingface\.co/", x, re.I) and "/datasets" not in x.lower()]
        J = _J.get(g.get("journal")) or {}
        note = (e.get("note") or "") + " " + (v.get("note") or "")
        dm = domain(g.get("title") or "", e.get("bench") or "", note)
        ws.append([i, g.get("title"), g.get("journal"), g.get("year"), dm,
                   e.get("bench"), e.get("size"), "; ".join(dl[:3]), e.get("sota"), e.get("metric"),
                   v.get("auto"), "; ".join(cl[:3]), f"https://doi.org/{r['doi']}", e.get("note"),
                   g.get("if") or J.get("jcr"), J.get("field") or "", dm,
                   art_type(r["doi"], g.get("title") or "", note), sota_status(e),
                   r["tier"], r["score"], v.get("open"), r.get("origin") or "", r.get("pdf", "")])
    ti = COLS.index("复刻优先级")
    TINT = {"S": "FCE4D6", "A": "FFF2CC", "B": "E2EFDA"}
    for rr in ws.iter_rows(min_row=2):
        t = rr[ti].value
        if t in TINT:
            fill = PatternFill("solid", fgColor=TINT[t])
            for c in rr:
                c.fill = fill
            if t == "S":
                rr[ti].font = Font(bold=True, color="C00000")
    ws.freeze_panes = "C2"
    m.style_sheet(ws, [5, 58, 24, 6, 14, 30, 40, 44, 60, 36, 9, 42, 40, 46,
                       8, 14, 16, 12, 18, 10, 8, 9, 12, 56])

    ws2 = wb.create_sheet("字段说明")
    ws2.append(["字段", "含义"])
    for k, vv in [
        ("PDF 文件名", "对应 pdfs/ 下的文件。命名沿用全文库的 期刊_年_标题_DOI尾号，"
                      "与排序无关——新增论文不会导致既有文件改名，所以不需要分批交付。"),
        ("复刻优先级", "S/A/B/C，由 可自动化(40)+基准公开(25)+SOTA标注(10)+期刊影响力"
                      "+链接与公开基准数+头顶空间(12) 加权，再扣人评/湿实验/硬件依赖"),
        ("复刻分", "上述加权总分，本表按此降序"),
        ("自动化等级", "A+ 可执行验证(单元测试/Pass@k) / A 有 ground truth 的客观指标 / "
                      "B 需 LLM-as-judge / C 需人评·湿实验·硬件"),
        ("基准公开", "是=有直达链接；部分=用公开基准但自建测试集受限；否=不公开"),
        ("来源", "PDF 取自哪个目录：fulltext=官方版；arxiv_隔离=只拿到 arXiv/作者稿；papers=既有语料"),
        ("Benchmark 规模", "从 PDF 正文抽取的样本量/任务数/队列规模，抽不到则标注说明，未做推测"),
        ("论文报告的 SOTA", "论文中可对比的具体数值，含对照方法名与数字；抽不到则明确标注"),
        ("评估指标 & 计算方式", "指标名与计算/划分协议，含需注意的评测陷阱"),
        ("复刻备注", "★数量表示推荐程度；⚠️标注需注意的限制（数据不公开、算力门槛、指标陷阱等）"),
    ]:
        ws2.append([k, vv])
    m.style_sheet(ws2, [20, 112])

    os.makedirs(OUT, exist_ok=True)
    xl = f"{OUT}/信息抽取.xlsx"
    wb.save(xl)
    print(f"  → {xl}  {len(rows)} 行 × {len(COLS)} 列")


if __name__ == "__main__":
    main(sys.argv[1:])
