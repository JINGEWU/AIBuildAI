# -*- coding: utf-8 -*-
"""S7 信息抽取 → papers/agent_papers_SOTA_evaluation.xlsx 的 14 字段结构

沿用既有表结构（config.EXTRACT_COLUMNS），只对第二轮「入选」论文抽取。
两类字段分开处理，避免把猜测当事实：
  机械可得（本步自动填）：期刊/年/领域/Benchmark 链接/代码链接/自动化等级/Paper link
  需读全文（dump→save 回填）：Benchmark 名称·规模、论文报告的 SOTA、评估指标计算方式、复刻备注

  import  把旧格式汇总表导入 extracted.json（让早期语料成为普通成员，不再特殊处理）
  dump    出待抽取批次（含 digest 证据）→ /tmp/x_<n>.txt
  save    回写抽取结果                  → data/extracted.json
  export  合并自动字段 + 回写字段出 xlsx

save 输入格式，每行以 `#i@doi` 开头，字段用 | 分隔：
  #0@10.xxxx/yyy|Benchmark名称|规模|SOTA数值|评估指标&计算方式|复刻备注
用法: python3 run.py info dump|save|export [--batch N] [--size N] [--file ...]
"""
import os, re, sys, json, argparse, collections
import config as C, common as m

EP = f"{C.DATA}/extracted.json"

def selected(pool=None):
    """第二轮入选，按复刻分降序 —— 先抽最该复刻的。pool 限定到某个交付批次。"""
    V = json.load(open(f"{C.DATA}/verdicts2.json", encoding="utf-8")) \
        if os.path.exists(f"{C.DATA}/verdicts2.json") else {}
    dg = {m.norm_doi(r["doi"]): r for r in m.read_jsonl(f"{C.DATA}/fulltext_digest.jsonl")}
    keep = {m.norm_doi(x) for x in json.load(open(pool, encoding="utf-8"))} if pool else None
    rows = []
    for d, v in V.items():
        if v.get("v") != "入选" or (keep is not None and d not in keep):
            continue
        g = dg.get(d, {})
        rows.append({"doi": d, "v": v, "d": g,
                     "score": m.repro_score(v, g, g.get("journal"))})
    rows.sort(key=lambda x: -x["score"])
    return rows

# 旧表（papers/agent_papers_SOTA_evaluation.xlsx）的列名 → extracted.json 的字段
LEGACY_MAP = {"bench": "Benchmark 名称", "size": "Benchmark 规模",
              "sota": "论文报告的 SOTA", "metric": "评估指标 & 计算方式",
              "note": "复刻备注"}


def import_xlsx(path, sheet=None, dry_run=False):
    """导入旧格式汇总表：抽取字段进 extracted.json，元数据回填 state，判定并入 verdicts2。

    早期那 54 篇的抽取结果一直存在自己的 xlsx 里，于是每次交付都要单独拼一次。
    导进来之后它们和其他论文走同一条路，S8 不必再有 --include-baseline 这种特例。
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hd = [str(x) if x is not None else "" for x in next(it)]
    rows = [dict(zip(hd, r)) for r in it]

    E = load_e()
    st = m.State("fetch_state_manual")
    VP2 = f"{C.DATA}/verdicts2.json"
    V = json.load(open(VP2, encoding="utf-8")) if os.path.exists(VP2) else {}
    bp = f"{C.DATA}/baseline_verdicts.json"
    B = json.load(open(bp, encoding="utf-8")) if os.path.exists(bp) else {}

    n_e = n_v = n_s = skip = 0
    for r in rows:
        mm = re.search(r"10\.\d{4,9}/\S+", str(r.get("Paper link") or ""))
        if not mm:
            skip += 1; continue
        d = m.norm_doi(mm.group(0))
        if d not in E:                       # 已有抽取（依据全文证据）的不覆盖
            E[d] = {k: str(r.get(col) or "").strip() for k, col in LEGACY_MAP.items()}
            n_e += 1
        if d not in V and d in B:            # verdicts2 的判定更可靠，只补不覆盖
            V[d] = B[d]; n_v += 1
        # 元数据回填：旧表用期刊缩写，统一成清单里的全名，否则 IF/领域/期刊加分都查不到
        j = str(r.get("期刊") or "").strip()
        cur = st.get(d) or {}
        if not cur.get("title") or not cur.get("journal") or not cur.get("year"):
            st[d] = {**cur, "ok": True, "src": cur.get("src") or "papers/已有",
                     "file": cur.get("file") or "",
                     "title": cur.get("title") or r.get("系统/论文") or r.get("系统·论文"),
                     "journal": cur.get("journal") or C.JOURNAL_ABBR.get(j, j),
                     "year": cur.get("year") or r.get("年")}
            n_s += 1
    if dry_run:
        m.log(f"dry-run：可导入抽取 {n_e}、判定 {n_v}、元数据 {n_s}（跳过无 DOI {skip}）")
        return
    st.flush()
    json.dump(E, open(EP, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    json.dump(V, open(VP2, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    m.log(f"导入抽取 {n_e} 条、判定 {n_v} 条、元数据 {n_s} 条（跳过无 DOI {skip}）；"
          f"extracted 累计 {len(E)}，verdicts2 累计 {len(V)}")
    m.log("下一步: run.py extract --refresh-meta  再  run.py build")


def load_e():
    return json.load(open(EP, encoding="utf-8")) if os.path.exists(EP) else {}

def _c(s, n):
    return re.sub(r"\s+", " ", str(s or ""))[:n]

def dump(batch, size, out, pool=None):
    rows = selected(pool)
    E = load_e()
    todo = [r for r in rows if r["doi"] not in E]
    seg = todo[batch * size:(batch + 1) * size]
    with open(out, "w", encoding="utf-8") as f:
        f.write(f"### 待抽取 {len(todo)} / 入选 {len(rows)}；本批 {batch*size}~"
                f"{batch*size+max(len(seg)-1,0)}\n")
        f.write("### 回写格式: #i@doi|Benchmark名称|规模|SOTA数值|评估指标&计算方式|复刻备注\n")
        for i, r in enumerate(seg):
            d, g, v = r["doi"], r["d"], r["v"]
            f.write(f"#{batch*size+i}@{d}|{m.repro_tier(r['score'])}{r['score']}|"
                    f"{_c(g.get('journal'),20)} {g.get('year')}|{_c(g.get('title'),95)}\n")
            f.write(f"  自动化={v.get('auto')} 基准公开={v.get('open')} SOTA={v.get('sota')}"
                    f" | 备注: {_c(v.get('note'),90)}\n")
            f.write(f"  链接: {'; '.join((g.get('links') or [])[:6]) or '无'}\n")
            f.write(f"  公开基准: {','.join(g.get('known_bench') or []) or '无'}"
                    f" | 自定义基准名: {','.join((g.get('bench_names') or [])[:6]) or '无'}\n")
            f.write(f"  指标: {','.join((g.get('metrics') or [])[:14]) or '无'}\n")
            for s in (g.get("avail") or [])[:1]:
                f.write(f"  DATA: {_c(s,340)}\n")
            for s in (g.get("sota") or [])[:3]:
                f.write(f"  SOTA: {_c(s,210)}\n")
            for s in (g.get("bench_sents") or [])[:2]:
                f.write(f"  BENCH: {_c(s,190)}\n")
    m.log(f"批 {batch}: {len(seg)} 条 → {out}（待抽取 {len(todo)}）")
    return out

def save(path):
    E = load_e()
    n, skipped = 0, 0
    for ln in open(path, encoding="utf-8"):
        ln = ln.rstrip("\n")
        if not ln.startswith("#") or "@" not in ln:
            continue
        head, _, rest = ln.partition("|")
        doi = m.norm_doi(head.split("@", 1)[1])
        c = rest.split("|")
        if not doi or len(c) < 4:
            skipped += 1
            continue
        E[doi] = {"bench": c[0].strip(), "size": c[1].strip(), "sota": c[2].strip(),
                  "metric": c[3].strip(), "note": c[4].strip() if len(c) > 4 else ""}
        n += 1
    json.dump(E, open(EP, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    m.log(f"回写 {n} 条" + (f"（跳过 {skipped} 行）" if skipped else "") + f"，累计 {len(E)}")
    return n

def export():
    rows = selected()
    E = load_e()
    wb = m.new_book()
    ws = wb.create_sheet("新增候选")
    ws.append(C.EXTRACT_COLUMNS + ["复刻优先级", "复刻分", "★待补抽取"])
    for i, r in enumerate(rows, 1):
        d, g, v = r["doi"], r["d"], r["v"]
        e = E.get(d, {})
        links = g.get("links") or []
        # openreview 是文献引用，不是数据/代码来源，两栏都不收
        links = [x for x in links if not re.match(r"openreview\.net", x, re.I)]
        DATA_RE = re.compile(r"(huggingface\.co/datasets|zenodo\.org|figshare\.com|kaggle\.com"
                             r"|physionet\.org|osf\.io|datadryad\.org|paperswithcode\.com)", re.I)
        CODE_RE = re.compile(r"(github\.com|gitlab\.com|codeocean\.com)", re.I)
        data_links = [x for x in links if DATA_RE.search(x)]
        code_links = [x for x in links if CODE_RE.search(x)]
        # HF 上非 datasets 路径的是模型权重，归到代码/权重栏
        code_links += [x for x in links
                       if re.match(r"huggingface\.co/", x, re.I) and "/datasets" not in x.lower()]
        ws.append([
            i,
            g.get("title"),
            g.get("journal"),
            g.get("year"),
            g.get("field") or "",
            e.get("bench") or ",".join((g.get("known_bench") or [])[:4]),
            e.get("size") or "",
            "; ".join(data_links[:3]),
            e.get("sota") or _c((g.get("sota") or [""])[0], 180),
            e.get("metric") or ",".join((g.get("metrics") or [])[:10]),
            v.get("auto"),
            "; ".join(code_links[:3]),
            f"https://doi.org/{d}",
            e.get("note") or v.get("note") or "",
            m.repro_tier(r["score"]), r["score"],
            "" if d in E else "待补",
        ])
    m.style_sheet(ws, [5, 58, 24, 6, 14, 28, 16, 46, 50, 34, 9, 40, 40, 44, 10, 8, 9])
    out = f"{C.ROOT}/信息抽取_新增候选.xlsx"
    wb.save(out)
    done = sum(1 for r in rows if r["doi"] in E)
    m.log(f"入选 {len(rows)} 篇，已完成读全文抽取 {done}，机械字段已全填 → {out}")
    m.log(f"提示: 未抽取的 {len(rows)-done} 篇，Benchmark 规模/SOTA 数值/指标计算方式"
          f" 目前只是 digest 里的原句，需 dump→save 补实")
    print("  优先级分布:", dict(collections.Counter(m.repro_tier(r["score"]) for r in rows)))

def main(argv):
    ap = argparse.ArgumentParser(prog="info")
    ap.add_argument("step", choices=["dump", "save", "export", "import"])
    ap.add_argument("--batch", type=int, default=0)
    ap.add_argument("--size", type=int, default=25)
    ap.add_argument("--out", default="/tmp/x_batch.txt")
    ap.add_argument("--file")
    ap.add_argument("--pool", help="只抽这个 json 里的 DOI 列表（交付批次）")
    ap.add_argument("--sheet", help="import 用：sheet 名，默认第一个")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args(argv)
    if a.step == "import":
        src = a.file or f"{C.PAPERS}/agent_papers_SOTA_evaluation.xlsx"
        import_xlsx(src, a.sheet, a.dry_run)
    elif a.step == "dump":
        dump(a.batch, a.size, a.out, a.pool)
    elif a.step == "save":
        if not a.file:
            m.die("save 需要 --file")
        save(a.file)
    else:
        export()
