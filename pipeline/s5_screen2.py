# -*- coding: utf-8 -*-
"""S5 第二轮筛选（全文）+ 复刻优先级 → 第二轮筛选结果.xlsx

  dump    把 digest 压成判读批次        → /tmp/w2_<n>.txt
  save    回写四维判定                  → data/verdicts2.json
  export  出 xlsx（含待手动下载清单）   → 第二轮筛选结果.xlsx

四维判定（每行 `#i@doi|判定|可自动化|基准公开|SOTA|备注`）：
  判定       入选 / 待定 / 排除
  可自动化   A+ 可执行验证 / A 客观指标 / B LLM-as-judge / C 人评·湿实验·硬件
  基准公开   是 / 部分 / 否
  SOTA       有 / 无
复刻优先级 S/A/B/C 由 common.repro_score 按 config 权重算出，S 档最值得先复刻。
用法: python3 run.py screen2 dump|save|export [--batch N] [--size N] [--file ...]
"""
import os, re, sys, json, argparse, collections
import config as C, common as m

VP = f"{C.DATA}/verdicts2.json"

def load_digest():
    """按 (复刻潜力, IF) 排序，保证判读顺序稳定 —— dump/save 都靠 DOI 定位，不怕顺序变。"""
    rows, seen = [], set()
    for r in m.read_jsonl(f"{C.DATA}/fulltext_digest.jsonl"):
        d = m.norm_doi(r.get("doi"))
        if not d or d in seen:
            continue
        seen.add(d)
        rows.append(r)
    def key(r):
        p = {"P1": 0, "P2": 1, "P3": 2, "P4": 3}.get(r.get("prio"), 9)
        return (p, -(r.get("if") or 0), r.get("doi") or "")
    rows.sort(key=key)
    return rows

# ───────────────────────── 判定规则（可执行版） ─────────────────────────
def sota_strength(d):
    """digest 里的 SOTA 证据强度：strong 有数值且有对比对象 / weak 只有数值 / none 只有字样。"""
    ss = d.get("sota") or []
    if not ss:
        return "none"
    num = [x for x in ss if C.RE_SOTA_NUM.search(x) and C.RE_SOTA_CLAIM.search(x)]
    if not num:
        return "none"
    return "strong" if any(C.RE_SOTA_BASELINE.search(x) for x in num) else "weak"

def has_data_evidence(d):
    """能不能真拿到测试集：有直达链接，或用了已知公开基准。"""
    return bool(d.get("links")) or bool(d.get("known_bench"))

def apply_rules(v, d):
    """按规则给出推荐判定 → (判定, 理由)。三道硬门槛 + 一条平局规则。"""
    reasons = []
    if v.get("open") in C.OPEN_BLOCK:
        reasons.append("测试集不公开")
    if v.get("auto") in C.AUTO_BLOCK:
        reasons.append("评估无法自动化(人评/湿实验/硬件)")
    if v.get("sota") in C.SOTA_BLOCK:
        reasons.append("未报告可对比 SOTA")
    if reasons:
        # 三条硬标准里挂了的，还有翻案余地就待定，否则排除
        salvage = has_data_evidence(d) and v.get("auto") not in C.AUTO_BLOCK
        return ("待定" if salvage else "排除"), "；".join(reasons)
    if v.get("open") == "部分" and C.PARTIAL_NEEDS_EVIDENCE and not has_data_evidence(d):
        return "待定", "仅声称部分公开，全文无任何数据/代码直达链接，需确认测试集可得性"
    return "入选", "三道门槛均通过"

def flags(v, d):
    """不自动改判、只提示复核的风险点。"""
    out = []
    if (d.get("n_hum") or 0) >= C.FLAG_HUM:
        out.append(f"人评词频{d['n_hum']}")
    if (d.get("n_wet") or 0) >= C.FLAG_WET:
        out.append(f"湿实验词频{d['n_wet']}")
    if (d.get("n_hw") or 0) >= C.FLAG_HW:
        out.append(f"硬件词频{d['n_hw']}")
    ss = sota_strength(d)
    if v.get("sota") == "有" and ss != "strong":
        out.append("SOTA证据弱(" + ("仅数值无对比对象" if ss == "weak" else "只有SOTA字样无数值") + ")")
    if v.get("open") == "是" and not (d.get("links") or []):
        out.append("标为完全公开但全文无直达链接")
    return out

def check():
    """拿规则复核已有判定，只报告不改写。"""
    rows = load_digest()
    V = load_v()
    dg = {m.norm_doi(r["doi"]): r for r in rows}
    dis, flagged = [], []
    agree = collections.Counter()
    for doi, v in V.items():
        d = dg.get(doi, {})
        rec, why = apply_rules(v, d)
        agree[(v.get("v"), rec)] += 1
        if rec != v.get("v"):
            dis.append((doi, v, rec, why, d))
        f = flags(v, d)
        if f:
            flagged.append((doi, v, f, d))
    m.log(f"复核 {len(V)} 条：规则与人工判定一致 "
          f"{sum(n for (a, b), n in agree.items() if a == b)}，不一致 {len(dis)}")
    print("\n  判定矩阵（行=人工，列=规则推荐）:")
    for a in ("入选", "待定", "排除"):
        print(f"    {a}: " + "  ".join(f"{b}{agree.get((a,b),0):4d}" for b in ("入选", "待定", "排除")))
    if dis:
        print(f"\n  不一致明细（前 20）:")
        for doi, v, rec, why, d in dis[:20]:
            print(f"    人工={v.get('v')} 规则={rec}  [{v.get('auto')}/{v.get('open')}/{v.get('sota')}]"
                  f"  {(d.get('title') or doi)[:52]}")
            print(f"        规则理由: {why}")
    print(f"\n  需复核的风险项 {len(flagged)} 条，按类型:")
    kinds = collections.Counter(k.split("(")[0].rstrip("0123456789") for _, _, fs, _ in flagged for k in fs)
    for k, n in kinds.most_common():
        print(f"    {k:28s} {n:4d}")
    sel_flag = [x for x in flagged if x[1].get("v") == "入选"]
    print(f"  其中已入选却带风险项的 {len(sel_flag)} 条（前 10）:")
    for doi, v, fs, d in sel_flag[:10]:
        print(f"    {(d.get('title') or doi)[:58]}")
        print(f"        {'; '.join(fs)}")
    return len(dis)

def load_v():
    return json.load(open(VP, encoding="utf-8")) if os.path.exists(VP) else {}

def _c(s, n):
    return re.sub(r"\s+", " ", str(s or ""))[:n]

# ───────────────────────── dump ─────────────────────────
def dump(batch, size, out):
    rows = load_digest()
    V = load_v()
    todo = [r for r in rows if m.norm_doi(r["doi"]) not in V]
    seg = todo[batch * size:(batch + 1) * size]
    with open(out, "w", encoding="utf-8") as f:
        f.write(f"### 未判 {len(todo)}（全文已抽取 {len(rows)}）本批 {batch*size}~"
                f"{batch*size+max(len(seg)-1,0)}\n")
        for i, r in enumerate(seg):
            doi = m.norm_doi(r["doi"])
            f.write(f"#{batch*size+i}@{doi}|{r.get('prio')}|{_c(r.get('journal'),20)} "
                    f"{r.get('year')} IF{r.get('if')}|{_c(r.get('title'),95)}\n")
            if r.get("err"):
                f.write(f"  ⚠ {r['err']}\n")
                continue
            f.write(f"  链接({len(r.get('links') or [])}): "
                    f"{'; '.join((r.get('links') or [])[:5]) or '无'}\n")
            f.write(f"  公开基准: {','.join(r.get('known_bench') or []) or '无'}"
                    f" | 指标: {','.join(r.get('metrics') or [])[:200] or '无'}\n")
            f.write(f"  红旗: 湿{r.get('n_wet',0)} 硬{r.get('n_hw',0)} 人评{r.get('n_hum',0)}\n")
            for s in (r.get("avail") or [])[:1]:
                f.write(f"  DATA: {_c(s,300)}\n")
            for s in (r.get("code_av") or [])[:1]:
                f.write(f"  CODE: {_c(s,200)}\n")
            for s in (r.get("sota") or [])[:2]:
                f.write(f"  SOTA: {_c(s,190)}\n")
            for s in (r.get("hum_ex") or [])[:1]:
                f.write(f"  人评例: {_c(s,130)}\n")
            for s in (r.get("wet_ex") or [])[:1]:
                f.write(f"  湿实验例: {_c(s,130)}\n")
    m.log(f"批 {batch}: {len(seg)} 条 → {out}（未判总数 {len(todo)}）")
    return out

# ───────────────────────── save ─────────────────────────
def save(path):
    V = load_v()
    n, skipped = 0, 0
    for ln in open(path, encoding="utf-8"):
        ln = ln.rstrip("\n")
        if not ln.startswith("#") or "@" not in ln:
            continue
        head, _, rest = ln.partition("|")
        doi = m.norm_doi(head.split("@", 1)[1])
        c = rest.split("|")
        if not doi or len(c) < 4:
            skipped += 1
            continue
        V[doi] = {"v": c[0].strip(), "auto": c[1].strip(), "open": c[2].strip(),
                  "sota": c[3].strip(), "note": c[4].strip() if len(c) > 4 else ""}
        n += 1
    json.dump(V, open(VP, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    m.log(f"回写 {n} 条" + (f"（跳过格式不全 {skipped} 行）" if skipped else "") + f"，累计 {len(V)}")
    print("  判定:", dict(collections.Counter(v["v"] for v in V.values())))
    print("  自动化:", dict(collections.Counter(v.get("auto") for v in V.values())))
    return n

# ───────────────────────── export ─────────────────────────
_AUDIT = None
def audit_map():
    """DOI → 审计判定。以 audit 读 PDF 内容的结论为准，比抓取来源可靠：
    有 34 篇是从 unpaywall/publisher 抓到的，但内容其实是 arXiv 版。"""
    global _AUDIT
    if _AUDIT is None:
        _AUDIT = {}
        p = f"{C.DATA}/cands.json"
        if os.path.exists(p):
            for r in json.load(open(p, encoding="utf-8")):
                if r.get("doi"):
                    _AUDIT[m.norm_doi(r["doi"])] = r.get("kind") or "非官方"
    return _AUDIT

KIND_LABEL = {"arxiv": "arXiv版·需官方版", "preprint": "作者稿·需官方版",
              "preprint?": "疑似非官方·需核对", "invalid": "文件损坏·需重下"}

def ft_version(doi, st):
    """(版本标签, 是否需手动补官方版)。"""
    k = audit_map().get(doi)
    if k:
        return KIND_LABEL.get(k, f"非官方({k})·需官方版"), True
    src = (st.get(doi) or {}).get("src", "")
    if src == "arxiv":
        return "arXiv版·需官方版", True
    if src in ("papers/已有", "已存在"):
        return "已有", False
    if not src:
        return "未获取", True
    return f"官方({src})", False

def export():
    from openpyxl.styles import Font, PatternFill
    rows = load_digest()
    V = load_v()
    st = m.load_shards("fetch_state")
    dg = {m.norm_doi(r["doi"]): r for r in rows}
    base = set()
    bp = f"{C.DATA}/baseline_dois.json"
    if os.path.exists(bp):
        base = {m.norm_doi(x) for x in json.load(open(bp, encoding="utf-8"))}
    s = m.xlsx_styles()

    cols = ["#", "复刻优先级", "复刻分", "判定", "可自动化", "基准公开", "SOTA标注", "备注",
            "★重复", "全文版本", "论文标题", "期刊", "年", "IF", "公开基准", "代码/数据链接",
            "评估指标", "湿实验", "硬件", "人评", "DOI", "DOI 链接"]

    def row_of(i, doi, v):
        d = dg.get(doi, {})
        sc = m.repro_score(v, d, d.get("journal"))
        ver, _ = ft_version(doi, st)
        return [i, m.repro_tier(sc), sc, v.get("v"), v.get("auto"), v.get("open"),
                v.get("sota"), v.get("note"), "★重复" if doi in base else "", ver,
                d.get("title"), d.get("journal"), d.get("year"), d.get("if"),
                ",".join(d.get("known_bench") or []), "; ".join((d.get("links") or [])[:5]),
                ",".join((d.get("metrics") or [])[:12]),
                d.get("n_wet"), d.get("n_hw"), d.get("n_hum"),
                doi, f"https://doi.org/{doi}"]

    wb = m.new_book()
    by = collections.defaultdict(list)
    for doi, v in V.items():
        by[v.get("v")].append((doi, v))
    for name in ("入选", "待定", "排除"):
        items = by.get(name, [])
        items.sort(key=lambda x: -m.repro_score(x[1], dg.get(x[0], {}),
                                                (dg.get(x[0]) or {}).get("journal")))
        ws = wb.create_sheet(f"{name}({len(items)})")
        ws.append(cols)
        for i, (doi, v) in enumerate(items, 1):
            ws.append(row_of(i, doi, v))
        for r in ws.iter_rows(min_row=2):
            if r[1].value == "S":
                r[1].font = Font(bold=True, color="C00000")
                fill = PatternFill("solid", fgColor=s["S_fill"])
                for c in r:
                    c.fill = fill
            elif r[1].value == "A":
                fill = PatternFill("solid", fgColor=s["A_fill"])
                for c in r:
                    c.fill = fill
        m.style_sheet(ws, [5, 10, 8, 7, 9, 9, 9, 40, 8, 18, 58, 24, 6, 7,
                           26, 44, 32, 7, 7, 7, 30, 40])

    # ── 统计 ──
    ws = wb.create_sheet("统计")
    ws.append(["项", "值", "说明"])
    # 自动化等级只统计入选——排除/待定的等级不参与对标决策
    ca = collections.Counter(v.get("auto") for v in V.values() if v.get("v") == "入选")
    okn = sum(1 for v in st.values() if v.get("ok"))
    fail = len(st) - okn
    arx = sum(1 for k, v in st.items() if v.get("ok") and ft_version(k, st)[1])
    for k, v, note in [
        ("全文已抽取", len(rows), ""),
        ("已判定", len(V), ""),
        ("入选", len(by.get("入选", [])), "有公开基准 + 可自动化评估 + 报告了 SOTA"),
        ("待定", len(by.get("待定", [])), "需要再确认基准可得性或评估方式"),
        ("排除", len(by.get("排除", [])), "测试集不公开 / 评估无法自动化 / 未报 SOTA"),
        ("入选中 A+ 可执行验证", ca.get("A+", 0), C.AUTOMATION_LEVELS["A+"]),
        ("入选中 A 客观指标", ca.get("A", 0), C.AUTOMATION_LEVELS["A"]),
        ("入选中 B LLM-as-judge", ca.get("B", 0), C.AUTOMATION_LEVELS["B"]),
        ("入选中 C 人评/湿实验/硬件", ca.get("C", 0), C.AUTOMATION_LEVELS["C"]),
        ("全文获取成功", okn, ""),
        ("全文获取失败", fail, "需机构权限手动补，见「待手动下载」页"),
        ("其中仅有 arXiv 版", arx, "你要求官方版，已并入「待手动下载」并标注②"),
    ]:
        ws.append([k, v, note])
    for tier in ("S", "A", "B", "C"):
        n = sum(1 for d2, v in V.items() if v.get("v") == "入选"
                and m.repro_tier(m.repro_score(v, dg.get(d2, {}),
                                               (dg.get(d2) or {}).get("journal"))) == tier)
        ws.append([f"—— 复刻优先级 {tier} 档", n, "" if tier != "S" else "最值得先做"])
    m.style_sheet(ws, [26, 10, 64])

    # ── 待手动下载 ──
    ws = wb.create_sheet("待手动下载")
    mcols = ["#", "原因", "复刻优先级", "判定", "筛选优先级", "论文标题", "期刊", "年", "IF",
             "DOI", "DOI 链接", "现有非官方文件名"]
    ws.append(mcols)
    mrows = []
    for doi, v in st.items():
        vv = V.get(doi, {})
        if not v.get("ok"):
            mrows.append(("① 完全未获取", "", "", v.get("prio"), v.get("title"),
                          v.get("journal"), v.get("year"), v.get("if"), doi, ""))
        elif ft_version(doi, st)[1]:
            if vv.get("v") == "排除":       # 已排除的不必再花权限去补
                continue
            tier = (m.repro_tier(m.repro_score(vv, dg.get(doi, {}),
                                               (dg.get(doi) or {}).get("journal")))
                    if vv.get("v") in ("入选", "待定") else "")
            lab, _ = ft_version(doi, st)
            mrows.append((f"② {lab}", tier, vv.get("v", ""), v.get("prio"),
                          v.get("title"), v.get("journal"), v.get("year"), v.get("if"),
                          doi, v.get("file") or ""))
    TO = {"S": 0, "A": 1, "B": 2, "C": 3, "": 4}
    PO = {"P1": 0, "P2": 1, "P3": 2, "P4": 3}
    mrows.sort(key=lambda x: (TO.get(x[1], 4), x[0], PO.get(x[3], 9), x[6] or ""))
    for i, (why, tier, vd, prio, ti, jn, yr, iff, doi, fn) in enumerate(mrows, 1):
        ws.append([i, why, tier, vd, prio, ti, jn, yr, iff, doi,
                   f"https://doi.org/{doi}", fn])
    for r in ws.iter_rows(min_row=2):
        if r[2].value == "S":
            r[2].font = Font(bold=True, color="C00000")
            fill = PatternFill("solid", fgColor=s["S_fill"])
            for c in r:
                c.fill = fill
        elif r[2].value == "A":
            fill = PatternFill("solid", fgColor=s["A_fill"])
            for c in r:
                c.fill = fill
    m.style_sheet(ws, [5, 24, 9, 7, 9, 58, 24, 6, 7, 30, 42, 52])

    # ── 判定标准 ──
    ws = wb.create_sheet("判定标准")
    ws.append(["字段", "取值", "含义"])
    for lv, desc in C.AUTOMATION_LEVELS.items():
        ws.append(["可自动化" if lv == "A+" else "", lv, desc])
    for r in [("基准公开", "是", "有 GitHub / HuggingFace / Zenodo / figshare 等直达链接"),
              ("", "部分", "使用 ImageNet/MIMIC/TCGA 等公开基准，但自建测试集受限或需申请"),
              ("", "否", "测试集不公开、需授权申请、或明确声明不释出"),
              ("SOTA标注", "有", "文中报告了可对比的 SOTA 数值"),
              ("", "无", "未报告可对比基线成绩"),
              ("红旗", "湿/硬/人评", "全文中相关词出现次数，用于快速定位自动化风险"),
              ("复刻优先级", "S/A/B/C",
               f"可自动化({max(C.W_AUTO.values())}) + 基准公开({max(C.W_OPEN.values())}) "
               f"+ SOTA({max(C.W_SOTA.values())}) + 期刊影响力 + 链接/基准数 "
               f"+ 头顶空间大({C.W_HEADROOM})，再扣人评/湿实验/硬件依赖"),
              ("复刻分", "0-100", "上述加权总分，各页按此降序")]:
        ws.append(list(r))
    m.style_sheet(ws, [16, 14, 92])

    wb.save(C.XLSX_ROUND2)
    m.log(f"入选{len(by.get('入选',[]))} 待定{len(by.get('待定',[]))} 排除{len(by.get('排除',[]))}"
          f" | A+{ca.get('A+',0)} A{ca.get('A',0)} B{ca.get('B',0)} C{ca.get('C',0)}")
    m.log(f"待手动下载 {len(mrows)} 篇 → {C.XLSX_ROUND2}")

def main(argv):
    ap = argparse.ArgumentParser(prog="screen2")
    ap.add_argument("step", choices=["dump", "save", "export", "check"])
    ap.add_argument("--batch", type=int, default=0)
    ap.add_argument("--size", type=int, default=45)
    ap.add_argument("--out", default="/tmp/w2_batch.txt")
    ap.add_argument("--file")
    a = ap.parse_args(argv)
    if a.step == "check":
        check()
    elif a.step == "dump":
        dump(a.batch, a.size, a.out)
    elif a.step == "save":
        if not a.file:
            m.die("save 需要 --file")
        save(a.file)
    else:
        export()
