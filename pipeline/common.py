# -*- coding: utf-8 -*-
"""共享工具：限速 HTTP、断点续传 state、DOI 规范化、分片、日志、xlsx 样式。"""
import os, re, sys, json, time, gzip, io, ssl, hashlib, urllib.request, urllib.parse, urllib.error
import config as C

# ───────────────────────── 日志 ─────────────────────────
_T0 = time.time()
def log(*a):
    m = int(time.time() - _T0) // 60
    s = int(time.time() - _T0) % 60
    print(f"[{m:02d}:{s:02d}]", *a, flush=True)

def die(msg, code=1):
    print(f"错误: {msg}", file=sys.stderr, flush=True)
    sys.exit(code)

# ───────────────────────── HTTP ─────────────────────────
_CTX = ssl.create_default_context()
_CTX.check_hostname = False
_CTX.verify_mode = ssl.CERT_NONE
_last = {}

def _wait(bucket):
    """按 config.RATE 做 per-source 限速。"""
    gap = C.RATE.get(bucket, 1.0)
    prev = _last.get(bucket, 0.0)
    d = gap - (time.time() - prev)
    if d > 0:
        time.sleep(d)
    _last[bucket] = time.time()

def http(url, bucket="default", data=None, headers=None, timeout=60,
         retries=3, backoff=2.0, binary=False, opener=None):
    """限速 + 重试 + gzip 解压。返回 str（或 bytes if binary）；失败返回 None。

    429/503 会按 Retry-After 或指数退避重试；4xx（除 429）不重试。
    """
    h = {"User-Agent": C.UA, "Accept-Encoding": "gzip"}
    if headers:
        h.update(headers)
    for attempt in range(retries):
        _wait(bucket)
        try:
            req = urllib.request.Request(url, data=data, headers=h)
            op = opener.open if opener else (lambda r, **k: urllib.request.urlopen(r, context=_CTX, **k))
            with op(req, timeout=timeout) as r:
                raw = r.read()
                if r.headers.get("Content-Encoding") == "gzip":
                    raw = gzip.decompress(raw)
                return raw if binary else raw.decode("utf-8", "replace")
        except urllib.error.HTTPError as e:
            if e.code in (429, 503) and attempt < retries - 1:
                ra = e.headers.get("Retry-After")
                # Retry-After 可能是「几万秒」（配额耗尽），那种情况没必要等
                nap = min(float(ra), 60) if (ra or "").isdigit() else backoff ** (attempt + 1)
                log(f"  {e.code} 退避 {nap:.0f}s  {url[:70]}")
                time.sleep(nap)
                continue
            return None
        except Exception:
            if attempt < retries - 1:
                time.sleep(backoff ** (attempt + 1))
                continue
            return None
    return None

def http_json(url, **kw):
    t = http(url, **kw)
    if not t:
        return None
    try:
        return json.loads(t)
    except Exception:
        return None

# ───────────────────────── DOI ─────────────────────────
def norm_doi(d):
    """统一小写、去 URL 前缀、去尾部标点。全流程以此为主键。"""
    if not d:
        return ""
    d = str(d).strip().lower()
    d = re.sub(r"^(https?://)?(dx\.)?doi\.org/", "", d)
    d = re.sub(r"^doi:\s*", "", d)
    return d.rstrip(" .,;)")

def doi_in_text(t):
    m = re.search(r"10\.\d{4,5}/[^\s,;)\"']{4,60}", t or "")
    return norm_doi(m.group(0)) if m else None

def safe_name(s, n=58):
    """把标题压成安全文件名片段。"""
    s = re.sub(r"<[^>]+>", "", s or "")
    s = re.sub(r"[^\w一-鿿 .\-]", "_", s)
    return re.sub(r"\s+", " ", s).strip()[:n]

# ───────────────────────── state（断点续传） ─────────────────────────
class State:
    """DOI → 记录 的 JSON 状态文件，支持分片与周期性落盘。

        st = State("fetch", shard=(0,5))
        if doi in st: continue
        st[doi] = {...}; st.flush()
    """
    def __init__(self, name, shard=None, autosave=25):
        tag = f"_{shard[0]}of{shard[1]}" if shard else ""
        self.path = f"{C.DATA}/{name}{tag}.json"
        self.shard = shard
        self.autosave = autosave
        self._n = 0
        self.d = {}
        if os.path.exists(self.path):
            try:
                self.d = json.load(open(self.path, encoding="utf-8"))
            except Exception:
                log(f"  state 损坏，重建: {self.path}")
    def __contains__(self, k): return norm_doi(k) in self.d
    def __getitem__(self, k): return self.d[norm_doi(k)]
    def __setitem__(self, k, v):
        self.d[norm_doi(k)] = v
        self._n += 1
        if self._n >= self.autosave:
            self.flush()
    def get(self, k, default=None): return self.d.get(norm_doi(k), default)
    def __len__(self): return len(self.d)
    def items(self): return self.d.items()
    def flush(self):
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        tmp = self.path + ".tmp"
        json.dump(self.d, open(tmp, "w", encoding="utf-8"), ensure_ascii=False)
        os.replace(tmp, self.path)   # 原子替换，避免中断留下半个文件
        self._n = 0

def load_shards(prefix):
    """把 data/{prefix}*.json 的所有分片合成一个 dict。"""
    out = {}
    if not os.path.isdir(C.DATA):
        return out
    for f in sorted(os.listdir(C.DATA)):
        if f.startswith(prefix) and f.endswith(".json"):
            try:
                out.update(json.load(open(f"{C.DATA}/{f}", encoding="utf-8")))
            except Exception:
                pass
    return out

def mine(key, shard):
    """分片归属判断：稳定哈希取模，保证多次运行分配一致。"""
    if not shard:
        return True
    i, n = shard
    return int(hashlib.md5(str(key).encode()).hexdigest(), 16) % n == i

def parse_shard(s):
    if not s:
        return None
    m = re.match(r"^(\d+)/(\d+)$", s)
    if not m:
        die(f"--shard 格式应为 i/N，收到 {s!r}")
    i, n = int(m.group(1)), int(m.group(2))
    if not (0 <= i < n):
        die(f"--shard 越界: {s}")
    return (i, n)

# ───────────────────────── jsonl ─────────────────────────
def read_jsonl(path, quiet=False):
    if not os.path.exists(path):
        if not quiet:
            log(f"  缺少 {path}")
        return []
    out = []
    for ln in open(path, encoding="utf-8"):
        ln = ln.strip()
        if ln:
            try:
                out.append(json.loads(ln))
            except Exception:
                pass
    return out

def write_jsonl(path, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    return len(rows)

def iter_jsonl(path):
    """流式读取，works_merged.jsonl 有 500MB，别整个 load 进内存。"""
    if not os.path.exists(path):
        return
    for ln in open(path, encoding="utf-8"):
        ln = ln.strip()
        if ln:
            try:
                yield json.loads(ln)
            except Exception:
                pass

# ───────────────────────── 全文文件索引 ─────────────────────────
# 三处（S4 抽取、S8 交付、pdftools 取证）以前各写一遍，目录集和优先级还不一样。
# 机制收在这里，但目录集由调用方显式传入 —— 各阶段该看哪些目录是有意的差异，不该是巧合：
#   S4        只认官方库（+可选 papers/），隔离区不进 digest
#   S8        要区分来源以便在交付表里标注 arXiv 版
#   pdftools  取证时 arXiv 版与官方版等价，全都要
FT_DIRS = ("fulltext", "fulltext_xml")
ALL_DIRS = FT_DIRS + ("papers", "arxiv_隔离")
_EXT = (".pdf", ".xml")

def file_index(dirs, with_origin=False, exts=_EXT):
    """文件名 → 路径（with_origin 时为 (路径, 来源目录)）。dirs 的顺序即优先级，先到为准。"""
    out = {}
    for d in dirs:
        p = f"{C.ROOT}/{d}"
        if not os.path.isdir(p):
            continue
        for f in sorted(os.listdir(p)):
            if f.startswith(".") or not f.lower().endswith(exts):
                continue
            out.setdefault(f, (f"{p}/{f}", d) if with_origin else f"{p}/{f}")
    return out

def fname(rec):
    """全文库统一命名：期刊缩写_年_标题_DOI尾号。S3 抓取与 S3 归档共用，两边必须一致。"""
    j = re.sub(r"[^A-Za-z]", "", (rec.get("journal") or "X"))[:22]
    return (f"{j}_{rec.get('year') or 'NA'}_{safe_name(rec.get('title'))}_"
            f"{norm_doi(rec['doi']).split('/')[-1].replace('/', '_')}")

# ───────────────────────── PDF / XML ─────────────────────────
_MUTED = False
def _silence_mupdf():
    """MuPDF 对轻微损坏的 PDF 会往 stdout 狂刷 format error，冲掉抽取结果；不影响取文本，静音。"""
    global _MUTED
    if not _MUTED:
        try:
            import fitz
            fitz.TOOLS.mupdf_display_errors(False)
        except Exception:
            pass
        _MUTED = True

def pdf_text(path, pages=None):
    import fitz
    _silence_mupdf()
    d = fitz.open(path)
    try:
        rng = range(d.page_count) if pages is None else range(min(pages, d.page_count))
        return "\n".join(d[i].get_text() for i in rng), d.page_count
    finally:
        d.close()

def xml_text(path):
    """JATS XML → 正文纯文本（剥掉参考文献、图表、公式）。"""
    t = open(path, encoding="utf-8", errors="replace").read()
    t = re.sub(r"<(ref-list|back|table-wrap|fig|inline-formula|disp-formula)[^>]*>.*?</\1>",
               " ", t, flags=re.S | re.I)
    t = re.sub(r"<[^>]+>", " ", t)
    t = (t.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
          .replace("&#x000a0;", " "))
    return re.sub(r"\s+", " ", t)

def sentences(t):
    t = re.sub(r"\s+", " ", t or "")
    return re.split(r"(?<=[.!?])\s+(?=[A-Z(])", t)

def is_real_pdf(path):
    """判断下载到的是真 PDF 还是登录页/HTML。返回 (ok, 原因)。"""
    try:
        if open(path, "rb").read(5) != b"%PDF-":
            return False, "非PDF(HTML/登录页)"
        if os.path.getsize(path) < C.MIN_PDF_BYTES:
            return False, "过小"
        return True, None
    except Exception as e:
        return False, f"读取失败:{e}"

def pdf_version(path):
    """审计 PDF 是官方版还是 arXiv/作者稿。返回 (版本, 原因, 页数)。"""
    ok, why = is_real_pdf(path)
    if not ok:
        return "invalid", why, 0
    try:
        t, n = pdf_text(path, pages=1)
    except Exception as e:
        return "invalid", f"解析失败:{e}", 0
    if C.RE_ARXIV_STAMP.search(t):
        return "arxiv", "含 arXiv 角标", n
    if C.RE_AUTHOR_LATEX.search(t):
        return "preprint", "作者 LaTeX 稿", n
    if n < C.MIN_PDF_PAGES:
        return "invalid", f"仅{n}页", n
    if not C.RE_PUBLISHER_STAMP.search(t):
        return "preprint?", "无出版社标记", n
    return "official", None, n

# ───────────────────────── xlsx 样式 ─────────────────────────
def xlsx_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="D9D9D9")
    return {
        "head_font": Font(bold=True, color="FFFFFF"),
        "head_fill": PatternFill("solid", fgColor="4472C4"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "wrap": Alignment(wrap_text=True, vertical="top"),
        "S_font": Font(bold=True, color="C00000"),
        "S_fill": "FCE4D6", "A_fill": "FFF2CC",
    }

def style_sheet(ws, widths, freeze="A2"):
    """表头样式 + 列宽 + 冻结首行 + 自动筛选。"""
    from openpyxl.utils import get_column_letter
    s = xlsx_styles()
    for c in ws[1]:
        c.font, c.fill = s["head_font"], s["head_fill"]
        c.alignment = s["wrap"]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    return ws

def new_book():
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    return wb

# ───────────────────────── 复刻优先级评分 ─────────────────────────
def repro_score(verdict, digest, journal=None):
    """按 config 权重算复刻分。verdict: {auto, open, sota, note}; digest: {links, known, n_wet...}"""
    v, d = verdict or {}, digest or {}
    s = 0.0
    s += C.W_AUTO.get(v.get("auto"), 0)
    s += C.W_OPEN.get(v.get("open"), 0)
    s += C.W_SOTA.get(v.get("sota"), 0)
    s += C.TOP_JOURNAL_BONUS.get(journal or d.get("journal"), C.TOP_JOURNAL_DEFAULT)
    s += min(C.W_LINK_CAP, len(d.get("links") or []) * C.W_LINK_EACH)
    s += min(C.W_KNOWN_CAP, len(d.get("known_bench") or d.get("known") or []) * C.W_KNOWN_EACH)
    if C.RE_HEADROOM.search(v.get("note") or ""):
        s += C.W_HEADROOM
    for k, (per, cap) in C.PENALTY.items():
        s -= min(cap, (d.get(f"n_{k}") or 0) * per)
    return round(s, 1)

def repro_tier(score):
    for name, cut in C.TIER_CUTS:
        if score >= cut:
            return name
    return "C"
