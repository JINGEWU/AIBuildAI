# -*- coding: utf-8 -*-
"""S11 全量索引：把 main/papers、batch2、batch3 汇成一张总表，实体文件一个不动。

为什么不真合并目录：batch2 与 batch3 各有一套 A001…/B001… 编号，56 组前缀重复。
平铺到同一目录虽不会覆盖（标题不同），但要重编号才排得清——那就破坏了同事在用的文件名。
所以只加索引层：新增「批次」与「PDF 路径」两列指向原位置。

main 那 54 篇原表只有 14 列，缺的 9 列在这里补齐：
  可直接算   JCR IF / 期刊领域（期刊清单查表）· 任务领域（s9 领域规则）
             SOTA可对比性（s9 三档判据作用在原有 SOTA 列上）· PDF 文件名（按 DOI 尾号匹配 papers/）
  需人工判   基准公开 / 自动化等级 / SOTA有无 → data/baseline_verdicts.json（依据原表 14 列内容回溯判定）
  据此算出   复刻分 / 复刻优先级 —— 用 TOP_JOURNAL_BONUS 按期刊全名取键，与 batch2/3 同一口径

用法: python3 run.py index [--out 全量索引.xlsx]
"""
import os, sys, re, json, argparse, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config as C, common as m
import s9_enrich as s9

MAIN_XLSX = f"{C.PAPERS}/agent_papers_SOTA_evaluation.xlsx"
COLS = C.EXTRACT_COLUMNS + ["JCR IF", "期刊领域", "任务领域", "文章类型",
                            "SOTA可对比性", "复刻优先级", "复刻分", "基准公开", "PDF 文件名"]
OUT_COLS = ["#", "批次"] + COLS[1:] + ["PDF 路径"]
# main 旧表的列名与新表有细微差别，逐一对齐
RENAME = {"系统/论文": "系统·论文", "Benchmark / 数据集链接": "Benchmark·数据集链接"}


def _read(path, sheet):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hd = [str(x) if x is not None else "" for x in next(it)]
    return [dict(zip(hd, r)) for r in it]


def _doi(row):
    mm = re.search(r"10\.\d{4,9}/\S+", str(row.get("Paper link") or ""))
    return m.norm_doi(mm.group(0)) if mm else None


def load_main():
    """main/papers 的 54 篇：对齐列名 + 补齐 9 个新列。"""
    rows = _read(MAIN_XLSX, "SOTA与评估方式")
    V = json.load(open(f"{C.DATA}/baseline_verdicts.json", encoding="utf-8"))
    D = {m.norm_doi(r["doi"]): r for r in m.read_jsonl(f"{C.DATA}/fulltext_digest.jsonl", True)}
    J = {j["name"]: j for j in C.journals()}
    files = [f for f in os.listdir(C.PAPERS) if f.lower().endswith(".pdf")] \
        if os.path.isdir(C.PAPERS) else []

    out, no_if = [], []
    for r in rows:
        d = _doi(r)
        r = {RENAME.get(k, k): v for k, v in r.items()}
        jfull = C.JOURNAL_ABBR.get(str(r.get("期刊") or "").strip(), str(r.get("期刊") or "").strip())
        j = J.get(jfull) or {}
        v = V.get(d) or {}
        g = D.get(d) or {}
        sc = m.repro_score(v, g, jfull)
        # 按 DOI 尾号回找 papers/ 里的文件名
        tail = (d or "").split("/")[-1].replace("/", "_")
        pdf = next((f for f in files if tail and os.path.splitext(f)[0].endswith("_" + tail)), "")
        if j.get("jcr") in (None, ""):
            no_if.append((jfull, d))
        r.update({
            "期刊": jfull,
            "JCR IF": j.get("jcr") if j.get("jcr") not in (None, "") else "清单未收录",
            "期刊领域": j.get("field") or "",
            "任务领域": s9.domain(r.get("系统·论文") or "", r.get("Benchmark 名称") or "",
                              r.get("复刻备注") or ""),
            "文章类型": ("纯基准评测" if d in C.BASELINE_PURE_BENCH else "原创方法"),
            "SOTA可对比性": s9.sota_status({"sota": r.get("论文报告的 SOTA")}),
            "复刻优先级": m.repro_tier(sc),
            "复刻分": sc,
            "基准公开": v.get("open") or "",
            "自动化等级": v.get("auto") or r.get("自动化等级"),
            "PDF 文件名": pdf,
            "_批次": "main", "_路径": f"papers/{pdf}" if pdf else "",
            "_原自动化": r.get("自动化等级"),
        })
        out.append(r)
    if no_if:
        cnt = collections.Counter(x[0] for x in no_if)
        m.log(f"期刊清单无 IF 的 {len(no_if)} 篇: {dict(cnt)}（标为「清单未收录」，不影响复刻分）")
    return out


def load_batch(name):
    """batchN 的表已是 23 列，原样取用。"""
    rows = _read(f"{C.ROOT}/{name}/信息抽取{name}.xlsx", f"{name}汇总")
    for r in rows:
        f = r.get("PDF 文件名") or ""
        r["_批次"] = name
        r["_路径"] = f"{name}/pdfs_{name}/{f}" if f else ""
        r["_原自动化"] = r.get("自动化等级")
    return rows


def main(argv):
    ap = argparse.ArgumentParser(prog="index")
    ap.add_argument("--out", default="全量索引.xlsx")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args(argv)

    parts = [("main", load_main())]
    for b in ("batch2", "batch3"):
        p = f"{C.ROOT}/{b}/信息抽取{b}.xlsx"
        if os.path.exists(p):
            parts.append((b, load_batch(b)))
        else:
            m.log(f"跳过 {b}（未找到 {p}）")
    rows = [r for _n, rs in parts for r in rs]

    # 交叉去重核查——三套本应互不重复，真撞了要显式报出来
    seen = collections.defaultdict(list)
    for r in rows:
        d = _doi(r)
        if d:
            seen[d].append(r["_批次"])
    dup = {d: b for d, b in seen.items() if len(b) > 1}
    for _n, rs in parts:
        print(f"  {_n:7s} {len(rs):>4d} 篇")
    print(f"  合计     {len(rows):>4d} 篇；唯一 DOI {len(seen)}；跨批重复 {len(dup)}")
    for d, b in list(dup.items())[:10]:
        print(f"     ⚠️ {d} 同时出现在 {b}")
    if a.dry_run:
        return

    # 按复刻分降序，缺分的排最后
    def key(r):
        try:
            return -float(r.get("复刻分") or -1)
        except (TypeError, ValueError):
            return 1e9
    rows.sort(key=key)

    from openpyxl.styles import Font, PatternFill
    wb = m.new_book(); ws = wb.create_sheet("全量索引")
    ws.append(OUT_COLS)
    for i, r in enumerate(rows, 1):
        ws.append([i, r["_批次"]] + [r.get(c) for c in COLS[1:]] + [r["_路径"]])
    TINT = {"S": "FCE4D6", "A": "FFF2CC", "B": "E2EFDA"}
    ti = OUT_COLS.index("复刻优先级")
    for rr in ws.iter_rows(min_row=2):
        t = rr[ti].value
        if t in TINT:
            for c in rr:
                c.fill = PatternFill("solid", fgColor=TINT[t])
            if t == "S":
                rr[ti].font = Font(bold=True, color="C00000")
    ws.freeze_panes = "C2"
    m.style_sheet(ws, [5, 8, 46, 24, 6, 20, 24, 30, 34, 60, 40, 10, 30, 30, 44,
                       10, 10, 14, 12, 16, 10, 8, 10, 52, 46])

    ws2 = wb.create_sheet("说明")
    ws2.append(["项", "说明"])
    for k, v in [
        ("用途", "main/papers、batch2、batch3 三套的统一入口。实体 PDF 一个未动，"
                 "本表只用「PDF 路径」列指向原位置。"),
        ("为什么不合并目录", "batch2 与 batch3 各有一套 A001…/B001… 编号，56 组前缀重复；"
                            "平铺需重编号，会破坏已在使用的文件名。"),
        ("批次", "main = papers/ 的 54 篇（早期语料）；batch2 = 254 篇；batch3 = 60 篇。三套互不重复。"),
        ("PDF 路径", "仓库相对路径，可直接定位文件。"),
        ("排序", "按复刻分降序，与各批次内部口径一致；# 列为本表行号，与各批次原表的 # 不对应。"),
        ("main 的 9 个新列", "原表只有 14 列。JCR IF/期刊领域来自期刊清单，任务领域与 SOTA可对比性"
                            "由规则作用在原有列上，基准公开/自动化等级/SOTA有无为逐篇回溯判定"
                            "（data/baseline_verdicts.json），复刻分据此算出。"),
        ("复刻分可比性", "复刻分的期刊项取自 TOP_JOURNAL_BONUS 按期刊全名查表，不依赖 JCR IF 数值，"
                        "因此三套的分数同一口径可直接比较。"),
        ("JCR IF「清单未收录」", "npj Digital Medicine、Nature Computational Science、Nature Health "
                                "在期刊清单里没有 IF（新刊或未纳入扫描），未做估算。不影响复刻分。"),
        ("自动化等级", "main 原表用 A(方法)/C(数据) 这类复合值，本表统一取主判据；原始复合值见 papers/ 原表。"),
    ]:
        ws2.append([k, v])
    m.style_sheet(ws2, [22, 112])

    out = f"{C.ROOT}/{a.out}"
    wb.save(out)
    print(f"\n→ {out}  {len(rows)} 行 × {len(OUT_COLS)} 列")
    for c in ("批次", "复刻优先级", "文章类型", "SOTA可对比性", "基准公开", "自动化等级"):
        i = OUT_COLS.index(c)
        print(f"  {c}: {dict(collections.Counter(str(r[i].value) for r in ws.iter_rows(min_row=2)))}")


if __name__ == "__main__":
    main(sys.argv[1:])
