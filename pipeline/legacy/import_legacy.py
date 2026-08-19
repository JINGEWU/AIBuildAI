# -*- coding: utf-8 -*-
"""复现工具：把旧格式汇总表导入 extracted.json / verdicts2.json。

papers/ 那 54 篇是早期人工收集的，抽取结果一直存在自己的 xlsx 里，
于是每次交付都要单独拼一次。导进来之后它们和其他论文走同一条路，
正式流程的 S8 不必再有 --include-baseline 这种特例。

这是一次性迁移，已经执行过（38 条抽取、29 条判定、54 条元数据）。
留在这里是为了可追溯与可重跑，不属于正式流程。

用法: python3 legacy/reproduce.py import-legacy [--file 旧表.xlsx] [--sheet 名] [--dry-run]
"""
import os, sys, re, json
sys.path[:0] = [os.path.dirname(os.path.abspath(__file__)),
                os.path.dirname(os.path.dirname(os.path.abspath(__file__)))]
import config as C, common as m
import tables as T

EP = f"{C.DATA}/extracted.json"


def load_e():
    return json.load(open(EP, encoding="utf-8")) if os.path.exists(EP) else {}


# 旧表（papers/agent_papers_SOTA_evaluation.xlsx）的列名 → extracted.json 的字段
LEGACY_MAP = {"bench": "Benchmark 名称", "size": "Benchmark 规模",
              "sota": "论文报告的 SOTA", "metric": "评估指标 & 计算方式",
              "note": "复刻备注"}


def import_xlsx(path, sheet=None, dry_run=False):
    """导入旧格式汇总表：抽取字段进 extracted.json，元数据回填 state，判定并入 verdicts2。

    早期那 54 篇的抽取结果一直存在自己的 xlsx 里，于是每次交付都要单独拼一次。
    导进来之后它们和其他论文走同一条路，S8 不必再有 --include-baseline 这种特例。
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hd = [str(x) if x is not None else "" for x in next(it)]
    rows = [dict(zip(hd, r)) for r in it]

    E = load_e()
    st = m.State("fetch_state_manual")
    VP2 = f"{C.DATA}/verdicts2.json"
    V = json.load(open(VP2, encoding="utf-8")) if os.path.exists(VP2) else {}
    bp = f"{C.DATA}/baseline_verdicts.json"
    B = json.load(open(bp, encoding="utf-8")) if os.path.exists(bp) else {}

    n_e = n_v = n_s = skip = 0
    for r in rows:
        mm = re.search(r"10\.\d{4,9}/\S+", str(r.get("Paper link") or ""))
        if not mm:
            skip += 1; continue
        d = m.norm_doi(mm.group(0))
        if d not in E:                       # 已有抽取（依据全文证据）的不覆盖
            E[d] = {k: str(r.get(col) or "").strip() for k, col in LEGACY_MAP.items()}
            n_e += 1
        if d not in V and d in B:            # verdicts2 的判定更可靠，只补不覆盖
            V[d] = B[d]; n_v += 1
        # 元数据回填：旧表用期刊缩写，统一成清单里的全名，否则 IF/领域/期刊加分都查不到
        j = str(r.get("期刊") or "").strip()
        cur = st.get(d) or {}
        if not cur.get("title") or not cur.get("journal") or not cur.get("year"):
            st[d] = {**cur, "ok": True, "src": cur.get("src") or "papers/已有",
                     "file": cur.get("file") or "",
                     "title": cur.get("title") or r.get("系统/论文") or r.get("系统·论文"),
                     "journal": cur.get("journal") or T.JOURNAL_ABBR.get(j, j),
                     "year": cur.get("year") or r.get("年")}
            n_s += 1
    if dry_run:
        m.log(f"dry-run：可导入抽取 {n_e}、判定 {n_v}、元数据 {n_s}（跳过无 DOI {skip}）")
        return
    st.flush()
    json.dump(E, open(EP, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    json.dump(V, open(VP2, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    m.log(f"导入抽取 {n_e} 条、判定 {n_v} 条、元数据 {n_s} 条（跳过无 DOI {skip}）；"
          f"extracted 累计 {len(E)}，verdicts2 累计 {len(V)}")
    m.log("下一步: run.py extract --refresh-meta  再  run.py build")
