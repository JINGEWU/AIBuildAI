#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""复现已发布的交付物 —— 独立于正式 pipeline。

正式 pipeline（`pipeline/run.py`）不分批次：一套判据、稳定命名、幂等，
未来有新论文跑一遍就是完整交付。这里放的是**只为复现历史交付**的东西，
它们带着两样正式流程已经不用的包袱：

  · `A001_标题.pdf` 这种按复刻分排序的档位编号 —— 加论文就要全部重命名，
    正是当初逼出「分批交付」的根源
  · main 那 54 篇的旧 14 列表结构

之所以留着：batch2 的 254 篇、batch3 的 60 篇已经发出去、同事在引用，
出问题时得能一字不差地重建出来对账。日常不要用这里的东西。

  reproduce.py batch2          重建 batch2/（254 篇，A001_ 编号）
  reproduce.py batch3          重建 batch3/（60 篇）
  reproduce.py import-legacy   把 main 的旧表导入 extracted.json（一次性迁移）
  reproduce.py verify          核对三套交付物与当前数据是否仍然自洽
"""
import os, sys, json, re, argparse, subprocess, collections

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path[:0] = [HERE, ROOT]
import config as C, common as m

# 各批次当初的构建参数，照此才能一字不差重建
RECIPES = {
    "batch2": dict(tiers="SAB", pool=None,
                   note="含与 main 重复的 16 篇剔除；剔除综述 1 篇、仅 XML 3 篇"),
    "batch3": dict(tiers="SAB", pool="batch3_selected.json",
                   note="来自 114 篇补齐下载；已排除与 batch2 重合的 8 篇"),
}


def _die_if_exists(batch):
    d = f"{C.ROOT}/{batch}/pdfs_{batch}"
    if os.path.isdir(d) and os.listdir(d):
        n = len([f for f in os.listdir(d) if f.lower().endswith(".pdf")])
        m.log(f"⚠️ {d} 已有 {n} 个 PDF。重建会按当前分数重新编号，"
              f"与已发布的文件名冲突。\n"
              f"   只想重出表格用 --xlsx-only；确认要作废旧名重建加 --force。")


def rebuild(batch, extra):
    r = RECIPES[batch]
    argv = ["--batch", batch, "--tiers", r["tiers"]]
    if r["pool"]:
        p = f"{C.DATA}/{r['pool']}"
        if not os.path.exists(p):
            m.die(f"缺 {p}（{batch} 的入选名单），无法复现")
        argv += ["--pool", p]
    argv += extra
    m.log(f"复现 {batch}：{r['note']}")
    _die_if_exists(batch)
    import build_batch
    return build_batch.main(argv)


def verify():
    """核对三套已发布交付物：行数、PDF 数、一一对应、跨批重复。"""
    import openpyxl
    sets, ok = {}, True
    specs = [("main", f"{C.PAPERS}/agent_papers_SOTA_evaluation.xlsx", "SOTA与评估方式", C.PAPERS),
             ("batch2", f"{C.ROOT}/batch2/信息抽取batch2.xlsx", "batch2汇总",
              f"{C.ROOT}/batch2/pdfs_batch2"),
             ("batch3", f"{C.ROOT}/batch3/信息抽取batch3.xlsx", "batch3汇总",
              f"{C.ROOT}/batch3/pdfs_batch3")]
    for name, xl, sh, pdir in specs:
        if not os.path.exists(xl):
            print(f"  · {name:7s} 表不存在，跳过"); continue
        wb = openpyxl.load_workbook(xl, read_only=True, data_only=True)
        ws = wb[sh] if sh in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hd = [str(x) if x is not None else "" for x in next(it)]
        rows = [dict(zip(hd, r)) for r in it]
        dois = set()
        for r in rows:
            mm = re.search(r"10\.\d{4,9}/\S+", str(r.get("Paper link") or ""))
            if mm:
                dois.add(m.norm_doi(mm.group(0)))
        sets[name] = dois
        have = {f for f in os.listdir(pdir) if f.lower().endswith(".pdf")} \
            if os.path.isdir(pdir) else set()
        col = "PDF 文件名"
        named = {r[col] for r in rows if r.get(col)} if col in hd else have
        extra, miss = have - named, named - have
        bad = (len(rows) != len(dois)) or extra or miss
        ok = ok and not bad
        print(f"  {'✗' if bad else '✓'} {name:7s} {len(rows):>3d} 行 / 唯一 DOI {len(dois):>3d} / "
              f"PDF {len(have):>3d}  多余 {len(extra)} 缺失 {len(miss)}")
        for f in list(extra)[:3]:
            print(f"        多余: {f[:70]}")
        for f in list(miss)[:3]:
            print(f"        缺失: {f[:70]}")
    names = list(sets)
    for i, a in enumerate(names):
        for b in names[i + 1:]:
            dup = sets[a] & sets[b]
            if dup:
                ok = False
                print(f"  ✗ {a} 与 {b} 重复 {len(dup)} 篇: {sorted(dup)[:3]}")
    if len(names) > 1:
        print(f"  并集 {len(set().union(*sets.values()))} 篇")
    print("  → " + ("三套自洽" if ok else "有不一致，见上"))
    return 0 if ok else 1


def main():
    ap = argparse.ArgumentParser(prog="reproduce", description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("what", choices=["batch2", "batch3", "import-legacy", "verify"])
    a, extra = ap.parse_known_args()
    if a.what in RECIPES:
        return rebuild(a.what, extra)
    if a.what == "import-legacy":
        import import_legacy
        ap2 = argparse.ArgumentParser(prog="import-legacy")
        ap2.add_argument("--file", default=f"{C.PAPERS}/agent_papers_SOTA_evaluation.xlsx")
        ap2.add_argument("--sheet")
        ap2.add_argument("--dry-run", action="store_true")
        b = ap2.parse_args(extra)
        return import_legacy.import_xlsx(b.file, b.sheet, b.dry_run)
    return verify()


if __name__ == "__main__":
    sys.exit(main() or 0)
