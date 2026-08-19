# -*- coding: utf-8 -*-
"""构建 batch2/ 交付文件夹：仅含已抽取的 S/A/B 档论文 PDF + 汇总 xlsx。"""
import os,sys,json,shutil,re,collections; sys.path.insert(0,os.path.dirname(os.path.abspath(__file__)))
import config as C, common as m

from s9_enrich import domain, art_type, sota_status, REVIEW

import argparse


def main(argv=None):
    ap = argparse.ArgumentParser(prog="build")
    ap.add_argument("--force", action="store_true",
                    help="允许重建已存在的交付目录（会按当前分数重新编号，作废旧文件名）")
    ap.add_argument("--xlsx-only", action="store_true", help="只重出 xlsx，不动 PDF")
    ap.add_argument("--batch", default="batch2", help="批次名，决定输出目录与文件名")
    ap.add_argument("--pool", help="只收这个 json 里的 DOI（该批次的入选名单）")
    ap.add_argument("--tiers", default="SAB", help="纳入的复刻档位，如 SAB 或 SABC")
    a = ap.parse_args(argv or [])
    OUT=f"{C.ROOT}/{a.batch}"
    PDFDIR=f"{OUT}/pdfs_{a.batch}"
    POOL=({m.norm_doi(x) for x in json.load(open(a.pool,encoding="utf-8"))} if a.pool else None)
    TIERS=tuple(a.tiers)
    BATCH=a.batch
    _J={j['name']:j for j in C.journals()}
    V=json.load(open(f"{C.DATA}/verdicts2.json"))
    E=json.load(open(f"{C.DATA}/extracted.json"))
    D={m.norm_doi(r['doi']):r for r in m.read_jsonl(f"{C.DATA}/fulltext_digest.jsonl")}
    st=m.load_shards("fetch_state")

    # 已抽取 且 入选 且 S/A/B 档
    # XML 全文（EPMC JATS）剥离了表格与图注，抽出的信息不全，本批不纳入
    # 与 papers/ 下既有 54 篇语料去重（main 分支已收录，batch2 不重复交付）
    _BASE=set()
    try:
        _BASE={m.norm_doi(x) for x in json.load(open(f"{C.DATA}/baseline_dois.json",encoding="utf-8"))}
    except Exception:
        pass

    def is_xml_source(doi):
        s=st.get(doi) or {}
        return s.get('src')=='epmc-xml' or (s.get('file') or '').lower().endswith('.xml')

    rows=[]; skipped_xml=[]; skipped_rev=[]; skipped_dup=[]
    for d,v in V.items():
        if v.get('v')!='入选' or d not in E: continue
        if POOL is not None and d not in POOL: continue
        g=D.get(d,{})
        sc=m.repro_score(v,g,g.get('journal')); t=m.repro_tier(sc)
        if t not in TIERS: continue
        if is_xml_source(d):
            skipped_xml.append((t,d,(g.get('title') or '')[:60])); continue
        if d in REVIEW:                      # 综述/复现报告，非原创研究，剔除
            skipped_rev.append((t,d,(g.get('title') or '')[:60])); continue
        if d in _BASE:                       # 与 main 的 papers/ 重复
            skipped_dup.append((t,d,(g.get('title') or '')[:60])); continue
        rows.append({'doi':d,'v':v,'g':g,'e':E[d],'score':sc,'tier':t})
    if skipped_dup:
        print(f"排除与 main/papers 重复 {len(skipped_dup)} 篇:")
        for t,d,ti in sorted(skipped_dup): print(f"   [{t}] {d}  {ti}")
    if skipped_rev:
        print(f"排除综述/复现报告 {len(skipped_rev)} 篇:")
        for t,d,ti in sorted(skipped_rev): print(f"   [{t}] {d}  {ti}")
    if skipped_xml:
        print(f"排除 XML 来源 {len(skipped_xml)} 篇（信息不全）:")
        for t,d,ti in sorted(skipped_xml): print(f"   [{t}] {d}  {ti}")
    rows.sort(key=lambda x:(-x['score']))

    # 文件来源索引（fulltext 优先，隔离区兜底）
    files={}
    for dr in ('fulltext','papers','arxiv_隔离'):
        p=f"{C.ROOT}/{dr}"
        if os.path.isdir(p):
            for f in os.listdir(p): files.setdefault(f,(f"{p}/{f}",dr))

    def safe(s,n=70):
        s=re.sub(r"[^\w一-鿿 .\-]", "_", re.sub(r"<[^>]+>","",str(s or "")))
        return re.sub(r"\s+"," ",s).strip()[:n].rstrip(". ")

    # 交付目录一旦发过（PDF 名字已进 xlsx、已推上远端），重新编号会让两边对不上。
    # 所以默认拒绝覆盖：要么 --xlsx-only 复用现有文件名，要么 --force 明确作废重来。
    existing = ({f for f in os.listdir(PDFDIR) if f.lower().endswith(('.pdf', '.xml'))}
                if os.path.isdir(PDFDIR) else set())
    if existing and not (a.force or a.xlsx_only):
        m.die(f"{PDFDIR} 已有 {len(existing)} 个文件。\n"
              f"  重建会按当前分数重新连续编号，与已交付/已推送的文件名冲突。\n"
              f"  --xlsx-only 只重出 xlsx（复用现有文件名）  |  --force 作废旧名彻底重建")

    # --xlsx-only：按 DOI 反查现有文件名，保持与已交付内容一致
    keep_by_doi = {}
    if a.xlsx_only and existing:
        for r in rows:
            s_ = st.get(r['doi']) or {}
            tail = r['doi'].split('/')[-1]
            _t = safe((r['g'].get('title') or ''))[:40]
            hit = next((f for f in existing if _t and _t in f), None)
            if hit:
                keep_by_doi[r['doi']] = hit

    os.makedirs(PDFDIR,exist_ok=True)
    tier_n=collections.Counter(); copied=[];missing=[]
    for i,r in enumerate(rows,1):
        tier_n[r['tier']]+=1
        idx=f"{r['tier']}{tier_n[r['tier']]:03d}"
        s=st.get(r['doi']) or {}
        src=files.get(s.get('file') or '')
        if not src:
            tail=r['doi'].split('/')[-1]
            src=next((v for k,v in files.items() if len(tail)>6 and tail[:20] in k),None)
        if not src:
            missing.append(r); r['pdf']=''; continue
        path,origin=src
        ext=os.path.splitext(path)[1].lower()
        name=keep_by_doi.get(r['doi']) or f"{idx}_{safe(r['g'].get('title'))}{ext}"
        if not a.xlsx_only:
            shutil.copyfile(path,f"{PDFDIR}/{name}")
        r['pdf']=name; r['origin']=origin
        copied.append(r)
    verb = "沿用" if a.xlsx_only else "复制"
    print(f"S/A/B 已抽取 {len(rows)} 篇；{verb} PDF {len(copied)}，缺文件 {len(missing)}")
    print("  档位:",dict(tier_n))
    print("  来源:",dict(collections.Counter(r.get('origin') for r in copied)))
    json.dump([{k:v for k,v in r.items() if k not in ('g','v','e')} for r in rows],
              open(f"{C.DATA}/{BATCH}_rows.json",'w',encoding='utf-8'),ensure_ascii=False)

    # ── 汇总 xlsx ──
    wb=m.new_book(); ws=wb.create_sheet(f"{BATCH}汇总")
    cols=C.EXTRACT_COLUMNS+["JCR IF","期刊领域","任务领域","文章类型","SOTA可对比性","复刻优先级","复刻分","基准公开","PDF 文件名"]
    ws.append(cols)
    DATA_RE=re.compile(r"(huggingface\.co/datasets|zenodo\.org|figshare\.com|kaggle\.com|physionet\.org|osf\.io|datadryad\.org|paperswithcode\.com)",re.I)
    CODE_RE=re.compile(r"(github\.com|gitlab\.com|codeocean\.com)",re.I)
    for i,r in enumerate(rows,1):
        g,v,e=r['g'],r['v'],r['e']
        lk=[x for x in (g.get('links') or []) if not re.match(r"openreview\.net",x,re.I)]
        dl=[x for x in lk if DATA_RE.search(x)]
        cl=[x for x in lk if CODE_RE.search(x)]+[x for x in lk if re.match(r"huggingface\.co/",x,re.I) and "/datasets" not in x.lower()]
        J=_J.get(g.get('journal')) or {}
        note=(e.get('note') or '')+' '+(v.get('note') or '')
        _dm=domain(g.get('title') or '', e.get('bench') or '', note)
        ws.append([i,g.get('title'),g.get('journal'),g.get('year'),
            _dm,
            e.get('bench'),e.get('size'),"; ".join(dl[:3]),e.get('sota'),e.get('metric'),
            v.get('auto'),"; ".join(cl[:3]),f"https://doi.org/{r['doi']}",e.get('note'),
            g.get('if') or J.get('jcr'), J.get('field') or '',
            _dm,
            art_type(r['doi'], g.get('title') or '', note),
            sota_status(e),
            r['tier'],r['score'],v.get('open'),r.get('pdf','')])
    from openpyxl.styles import Font,PatternFill
    for rr in ws.iter_rows(min_row=2):
        t=rr[19].value
        if t=='S':
            rr[19].font=Font(bold=True,color="C00000")
            f=PatternFill("solid",fgColor="FCE4D6")
            for c in rr: c.fill=f
        elif t=='A':
            f=PatternFill("solid",fgColor="FFF2CC")
            for c in rr: c.fill=f
    m.style_sheet(ws,[5,58,24,6,14,30,40,44,60,36,9,42,40,46,8,14,16,12,18,10,8,9,50])

    ws2=wb.create_sheet("字段说明")
    ws2.append(["字段","含义"])
    for _k,_v in [("复刻优先级","S/A/B/C 四档，由可自动化等级(40)+基准公开度(25)+SOTA标注(10)+期刊影响力+链接与公开基准数+头顶空间大(12)加权，再扣人评/湿实验/硬件依赖"),
     ("复刻分","上述加权总分，本表按此降序"),
     ("自动化等级","A+ 可执行验证(单元测试/Pass@k) / A 有ground truth的客观指标 / B 需LLM-as-judge / C 需人评·湿实验·硬件"),
     ("基准公开","是=有直达链接；部分=用公开基准但自建测试集受限；否=不公开(本表不含)"),
     ("Benchmark 规模","从 PDF 正文抽取的样本量/任务数/队列规模，抽不到则标注说明，未做推测"),
     ("论文报告的 SOTA","论文中可对比的具体数值，含对照方法名与数字；抽不到则明确标注"),
     ("评估指标 & 计算方式","指标名与计算/划分协议，含需注意的评测陷阱"),
     ("复刻备注","★数量表示推荐程度；⚠️标注需注意的限制（数据不公开、算力门槛、指标陷阱等）"),
     ("PDF 文件名",f"对应 pdfs_{BATCH}/ 目录下的文件，前缀为 档位+序号")]:
        ws2.append([_k,_v])
    m.style_sheet(ws2,[18,110])
    wb.save(f"{OUT}/信息抽取{BATCH}.xlsx")
    print(f"→ {OUT}/信息抽取{BATCH}.xlsx  {len(rows)} 行")



if __name__ == "__main__":
    main(sys.argv[1:])