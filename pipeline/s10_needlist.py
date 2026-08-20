# -*- coding: utf-8 -*-
"""S10 手动下载清单：列出仍需人工去机构网关取官方 PDF 的论文。

三类缺口，按补齐价值排序：
  ① 全文从未抓到 —— 信息盲区，可能藏着 S/A 档，优先补
  ② 只有 arXiv/作者稿 —— 判读结论有效，但交付需要官方版
  ③ 只有 EPMC XML —— JATS 剥离了表格与图注，Benchmark 规模与 SOTA 抽不全

输出 xlsx（含说明页）+ csv（便于批量喂下载脚本）。
"""
import sys, os, json, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config as C, common as m

TIER_ORDER = {"S": 0, "A": 1, "B": 2, "C": 3, "": 4}
REASONS = [
    ("① 全文未获取（尚未判读，不知是否合格）", 1,
     "抓取全部失败（IEEE 无金色 OA、订阅墙等），从未判读过",
     "优先补：这批是信息盲区，里面可能有 S/A 档"),
    ("② 仅有 arXiv/作者稿，需官方版", 2,
     "已判读，结论有效，但 PDF 非出版社官方版",
     "按复刻优先级补，S/A 优先"),
    ("③ 仅有 EPMC XML，表格被剥离信息不全", 3,
     "JATS XML 剥离了表格与图注，Benchmark 规模与 SOTA 数值抽不全",
     "补官方 PDF 后可重新抽取，已暂时排除出交付批次"),
]


def publisher(doi):
    return C.PUBLISHER.get(doi.split("/")[0], "其他")


def collect():
    V = json.load(open(f"{C.DATA}/verdicts2.json", encoding="utf-8"))
    D = {m.norm_doi(r["doi"]): r for r in m.read_jsonl(f"{C.DATA}/fulltext_digest.jsonl", True)}
    st = m.load_shards("fetch_state")
    ft = set(os.listdir(C.FULLTEXT)) if os.path.isdir(C.FULLTEXT) else set()
    qdir = f"{C.ROOT}/arxiv_隔离"
    q = set(os.listdir(qdir)) if os.path.isdir(qdir) else set()
    J = {j["name"]: j for j in C.journals()}

    rows = []
    for d, s in st.items():
        g = D.get(d, {}); v = V.get(d, {}); f = s.get("file") or ""
        tier = sel = ""
        if d in V and d in D:
            tier = m.repro_tier(m.repro_score(v, g, g.get("journal")))
            sel = v.get("v", "")
        # 已人工判定剔除的（会议摘要等），不再挂在待补清单上
        if (s.get("src") or "").startswith("剔除/"):
            continue
        # 分类：未获取 > 仅非官方 > 仅 XML；都不是则已齐备，跳过
        if not s.get("ok"):
            why, prio = REASONS[0][0], REASONS[0][1]
        elif f and f in q and f not in ft:
            why, prio = REASONS[1][0], REASONS[1][1]
        elif s.get("src") == "epmc-xml" or f.endswith(".xml"):
            why, prio = REASONS[2][0], REASONS[2][1]
        else:
            continue
        jn = g.get("journal") or s.get("journal") or ""
        rows.append({
            "why": why, "prio": prio, "tier": tier, "sel": sel,
            "title": g.get("title") or s.get("title") or "", "journal": jn,
            "if": g.get("if") or s.get("if") or (J.get(jn) or {}).get("jcr"),
            "doi": d, "pub": publisher(d), "file": f})
    rows.sort(key=lambda r: (r["prio"], TIER_ORDER.get(r["tier"], 4), -(r["if"] or 0)))
    return rows


def main(argv):
    import argparse
    ap = argparse.ArgumentParser(prog="needlist")
    ap.add_argument("--out", default="待手动下载_batch2", help="输出文件名（不含扩展名）")
    a = ap.parse_args(argv)

    base = a.out if os.path.isabs(a.out) else f"{C.ROOT}/{a.out}"
    rows = collect()
    print(f"需手动下载合计 {len(rows)} 篇")
    for k, n in collections.Counter(r["why"] for r in rows).most_common():
        print(f"  {k}  {n}")
    print("\n按出版社:", dict(collections.Counter(r["pub"] for r in rows).most_common()))
    print("按复刻优先级:", dict(collections.Counter(r["tier"] or "未判读" for r in rows).most_common()))
    print("按期刊 Top10:", collections.Counter(r["journal"] for r in rows).most_common(10))

    from openpyxl.styles import Font, PatternFill
    wb = m.new_book(); ws = wb.create_sheet("待手动下载")
    ws.append(["#", "原因", "复刻优先级", "二轮判定", "论文标题", "期刊", "JCR IF",
               "出版社", "DOI", "DOI 链接", "现有非官方文件名"])
    for i, r in enumerate(rows, 1):
        ws.append([i, r["why"], r["tier"], r["sel"], r["title"], r["journal"], r["if"],
                   r["pub"], r["doi"], f"https://doi.org/{r['doi']}", r["file"]])
    for rr in ws.iter_rows(min_row=2):
        t = rr[2].value
        if t == "S":
            rr[2].font = Font(bold=True, color="C00000")
            for c in rr:
                c.fill = PatternFill("solid", fgColor="FCE4D6")
        elif t == "A":
            for c in rr:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
    m.style_sheet(ws, [5, 34, 9, 8, 62, 26, 8, 10, 30, 42, 52])

    ws2 = wb.create_sheet("说明")
    ws2.append(["原因", "含义", "建议"])
    for why, _p, meaning, advice in REASONS:
        ws2.append([why.split("（")[0], meaning, advice])
    m.style_sheet(ws2, [24, 56, 44])

    xl = f"{base}.xlsx"; wb.save(xl); print(f"\n→ {xl}")
    csv = f"{base}.csv"
    with open(csv, "w", encoding="utf-8-sig") as fo:
        fo.write("原因,优先级,判定,标题,期刊,IF,出版社,DOI,链接\n")
        for r in rows:
            t = (r["title"] or "").replace('"', "'")
            fo.write(f'"{r["why"]}","{r["tier"]}","{r["sel"]}","{t}","{r["journal"]}",'
                     f'{r["if"] or ""},"{r["pub"]}","{r["doi"]}","https://doi.org/{r["doi"]}"\n')
    print(f"→ {csv}")


if __name__ == "__main__":
    main(sys.argv[1:])
